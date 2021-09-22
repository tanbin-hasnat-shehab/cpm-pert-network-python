import streamlit as st
from openpyxl import load_workbook

wb=load_workbook('tmp/data.xlsx')
sheet=wb.active
t1=st.text_input(f'{1}')
t2=st.text_input(f'{2}')
if st.button('ok'):
    sheet.cell(row=1,column=1).value=t1
    wb.save('/tmp/data.xlsx')
    
    f = open("tmp/demofile.txt", "w")
    f.writelines([t1,'\n',t2])
    f.close()
    f = open("tmp/demofile.txt", "r")
    st.write(f.read())
    
